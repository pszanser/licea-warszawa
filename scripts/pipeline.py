import argparse
import asyncio
import datetime
import difflib
import json
import logging
import os
import re
import time
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import requests
import yaml

from scripts.api_clients.googlemaps_api import (
    get_coordinates_for_addresses_batch,
    get_next_weekday_time,
    get_travel_times_batch,
)
from scripts.config.constants import ALL_SUBJECTS
from scripts.data_processing.load_minimum_points import load_min_points
from scripts.data_processing.get_data_pzo_omikron import (
    DEFAULT_BASE_URL as PZO_BASE_URL,
    DEFAULT_PUBLIC_CONTEXT as PZO_PUBLIC_CONTEXT,
    PzoOmikronClient,
    build_tables as build_pzo_tables,
    fetch_offer_snapshot,
    load_snapshot_files as load_pzo_snapshot_files,
    write_snapshot_files,
)
from scripts.data_processing.parser_perspektywy import (
    parse_ranking_perspektywy_html,
    parse_ranking_perspektywy_pdf,
)

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
RESULTS_DIR = BASE_DIR / "results"
SCRIPTS_DIR = BASE_DIR / "scripts"
SOURCE_CONFIG_FILE = SCRIPTS_DIR / "config" / "data_sources.yml"
PROJECT_CONFIG_FILE = SCRIPTS_DIR / "config" / "config.yml"
KODY_FILE = DATA_DIR / "reference" / "waw_kod_dzielnica.csv"
CZASY_DOJAZDU_FILE = RESULTS_DIR / "czasy_dojazdu.xlsx"
LEGACY_APP_FILE = RESULTS_DIR / "LO_Warszawa_2025_Warszawa_SL.xlsx"
THRESHOLD_COLUMNS = [
    "NazwaSzkoly",
    "OddzialNazwa",
    "Prog_min_klasa",
    "threshold_year",
    "threshold_kind",
    "threshold_priority",
    "threshold_label",
    "threshold_source",
    "SzkolaIdentyfikator",
    "SymbolOddzialu",
    "year",
    "admission_year",
    "school_year",
]

REPL = {
    r"liceum og[oó]lnokszta[łl]c[a-ząćęłńóśźż]*": "lo",
    r"im(?:\.|ienia)?": "",
    r"i\.?i\.?": "ii",
}


def load_yaml(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def project_config() -> dict[str, Any]:
    return load_yaml(PROJECT_CONFIG_FILE)


def source_config() -> dict[str, Any]:
    return load_yaml(SOURCE_CONFIG_FILE)


def resolve_path(path_value: str | Path) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else BASE_DIR / path


def normalize_name(name: str) -> str:
    if pd.isna(name):
        return ""
    x = str(name).lower()
    for pat, rep in REPL.items():
        x = re.sub(pat, rep, x)
    x = unicodedata.normalize("NFKD", x).encode("ascii", "ignore").decode()
    x = re.sub(r"[^a-z0-9 ]+", " ", x)
    x = re.sub(r"\s+", " ", x).strip()
    m = re.match(r"(?P<num>[ivxlcdm]+)\s+lo.*?(?P<patron>[a-z]+)$", x)
    if m:
        return f"{m['num']}_{m['patron']}"
    return x


def get_school_type(name: str) -> str:
    name = str(name).lower()
    if "technikum" in name:
        return "technikum"
    if "branżowa" in name or "branzowa" in name:
        return "branżowa"
    return "liceum"


def extract_class_type(class_name: str) -> str | None:
    if pd.isna(class_name):
        return None
    m = re.search(r"\[([^\]]+)\]", str(class_name))
    return m.group(1).strip() if m else None


SUBJECT_MATCH_ALIASES = {
    "matematyka": "mat",
    "mat": "mat",
    "język angielski": "ang",
    "jezyk angielski": "ang",
    "angielski": "ang",
    "ang": "ang",
    "biologia": "biol",
    "biol": "biol",
    "chemia": "chem",
    "chem": "chem",
    "fizyka": "fiz",
    "fiz": "fiz",
    "geografia": "geo",
    "geogr": "geo",
    "geo": "geo",
    "informatyka": "inf",
    "inf": "inf",
    "wiedza o społeczeństwie": "wos",
    "wiedza o spoleczenstwie": "wos",
    "wos": "wos",
    "historia": "his",
    "hist": "his",
    "his": "his",
    "język polski": "pol",
    "jezyk polski": "pol",
    "polski": "pol",
    "pol": "pol",
    "biznes": "biz",
    "biz": "biz",
    "język hiszpański": "hiszp",
    "jezyk hiszpanski": "hiszp",
    "hiszpański": "hiszp",
    "hiszpanski": "hiszp",
    "hiszp": "hiszp",
    "hisz": "hiszp",
    "język niemiecki": "niem",
    "jezyk niemiecki": "niem",
    "niemiecki": "niem",
    "niem": "niem",
    "język francuski": "franc",
    "jezyk francuski": "franc",
    "francuski": "franc",
    "franc": "franc",
    "fra": "franc",
    "język włoski": "wlos",
    "jezyk wloski": "wlos",
    "włoski": "wlos",
    "wloski": "wlos",
    "wlo": "wlos",
}

LANGUAGE_MATCH_ALIASES = {
    "język angielski": "ang",
    "jezyk angielski": "ang",
    "angielski": "ang",
    "ang": "ang",
    "gb": "ang",
    "język niemiecki": "niem",
    "jezyk niemiecki": "niem",
    "niemiecki": "niem",
    "niem": "niem",
    "de": "niem",
    "język hiszpański": "hiszp",
    "jezyk hiszpanski": "hiszp",
    "hiszpański": "hiszp",
    "hiszpanski": "hiszp",
    "hisz": "hiszp",
    "hiszp": "hiszp",
    "es": "hiszp",
    "język francuski": "franc",
    "jezyk francuski": "franc",
    "francuski": "franc",
    "franc": "franc",
    "fr": "franc",
    "fra": "franc",
    "język rosyjski": "ros",
    "jezyk rosyjski": "ros",
    "rosyjski": "ros",
    "ros": "ros",
    "ru": "ros",
    "język włoski": "wlos",
    "jezyk wloski": "wlos",
    "włoski": "wlos",
    "wloski": "wlos",
    "wlo": "wlos",
    "wł": "wlos",
    "wl": "wlos",
    "it": "wlos",
    "język łaciński": "lac",
    "jezyk lacinski": "lac",
    "łaciński": "lac",
    "lacinski": "lac",
    "łacina": "lac",
    "lacina": "lac",
    "lac": "lac",
    "język portugalski": "port",
    "jezyk portugalski": "port",
    "portugalski": "port",
    "port": "port",
    "por": "port",
}

LANGUAGE_DISPLAY_BY_TOKEN = {
    "ang": "angielski",
    "niem": "niemiecki",
    "hiszp": "hiszpański",
    "franc": "francuski",
    "ros": "rosyjski",
    "wlos": "włoski",
    "lac": "łacina",
    "port": "portugalski",
}
LANGUAGE_TOKEN_BY_DISPLAY = {
    display: token for token, display in LANGUAGE_DISPLAY_BY_TOKEN.items()
}
LANGUAGE_LEVELS = ("dwujęzyczny", "kontynuacja", "od podstaw", "bez oznaczenia")

CLASS_TYPE_ALIASES = {
    "ogolnodostepny": "O",
    "ogólnodostępny": "O",
    "dwujezyczny": "D",
    "dwujęzyczny": "D",
    "wstepny": "W",
    "wstępny": "W",
    "integracyjny": "I",
    "sportowy": "S",
    "mistrzostwa sportowego": "MS",
}

ORDINAL_TO_COLUMN = {
    "pierwszy": "Punktowany1",
    "drugi": "Punktowany2",
    "trzeci": "Punktowany3",
    "czwarty": "Punktowany4",
}


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def ascii_key(value: Any) -> str:
    text = safe_text(value).lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_code(value: Any) -> str:
    text = safe_text(value).upper()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", "", text)


def extract_class_code(name: Any, explicit_code: Any = None) -> str:
    explicit = compact_code(explicit_code)
    if explicit:
        return explicit
    text = safe_text(name).strip()
    if not text:
        return ""
    head = re.split(r"\s+-\s+|\s+\[|\s+\(", text, maxsplit=1)[0]
    return compact_code(head)


def class_code_base(value: Any) -> str:
    code = compact_code(value)
    match = re.match(r"^(1?[A-Z]+|1?\d+[A-Z]+)", code)
    return match.group(1) if match else code[:3]


def class_code_similarity(current_code: Any, old_code: Any) -> tuple[float, str]:
    current = compact_code(current_code)
    old = compact_code(old_code)
    if current and old and current == old:
        return 1.0, "code_exact"
    if current and old:
        current_core = current.removeprefix("1")
        old_core = old.removeprefix("1")
        if current_core == old_core:
            return 0.9, "code_core"
        if (current.startswith(old) or old.startswith(current)) and min(
            len(current), len(old)
        ) >= 2:
            return 0.75, "code_prefix"
    if class_code_base(current) and class_code_base(current) == class_code_base(old):
        return 0.55, "code_base"
    return 0.0, ""


def token_set_from_text(value: Any, aliases: dict[str, str]) -> tuple[str, ...]:
    text = safe_text(value)
    normalized = ascii_key(text)
    words = set(normalized.split())
    tokens: list[str] = []
    for source, token in sorted(aliases.items(), key=lambda item: -len(item[0])):
        source_key = ascii_key(source)
        if not source_key:
            continue
        if " " in source_key or len(source_key) > 3:
            if re.search(rf"\b{re.escape(source_key)}\b", normalized):
                tokens.append(token)
        elif source_key in words:
            tokens.append(token)
    return tuple(sorted(set(tokens)))


def unique_preserving_order(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def language_level_from_text(value: Any) -> str:
    raw = safe_text(value).lower()
    normalized = ascii_key(raw)
    if "poziom dwujezyczny" in normalized or "dwujezyczny" in normalized or "*d" in raw:
        return "dwujęzyczny"
    if "kontynuacja" in normalized or "(k)" in raw:
        return "kontynuacja"
    if "od podstaw" in normalized or "(p)" in raw:
        return "od podstaw"
    return "bez oznaczenia"


def normalize_language_name(value: Any) -> str:
    tokens = token_set_from_text(value, LANGUAGE_MATCH_ALIASES)
    if not tokens:
        return ""
    return LANGUAGE_DISPLAY_BY_TOKEN.get(tokens[0], "")


def split_language_options(value: Any) -> list[str]:
    text = safe_text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"\s*[,;]\s*", text) if part.strip()]


def parse_language_option(value: Any) -> tuple[str, str] | None:
    language = normalize_language_name(value)
    if not language:
        return None
    return language, language_level_from_text(value)


def dedupe_language_options(
    options: list[tuple[str, str]],
) -> tuple[tuple[str, str], ...]:
    seen = set()
    result = []
    for language, level in options:
        key = (language, level)
        if language and key not in seen:
            seen.add(key)
            result.append(key)
    return tuple(result)


def parse_language_option_list(value: Any) -> tuple[tuple[str, str], ...]:
    options = []
    for part in split_language_options(value):
        parsed = parse_language_option(part)
        if parsed:
            options.append(parsed)
    return dedupe_language_options(options)


def parse_legacy_language_slots(
    value: Any,
) -> tuple[
    tuple[tuple[str, str], ...],
    tuple[tuple[str, str], ...],
    tuple[tuple[str, str], ...],
]:
    text = safe_text(value)
    if not text:
        empty: tuple[tuple[str, str], ...] = tuple()
        return empty, empty, empty

    first_text = ""
    second_text = ""
    first_match = re.search(r"(?:^|\b)1\s*:\s*(.*?)(?=\b2\s*:|$)", text, re.I)
    second_match = re.search(r"(?:^|\b)2\s*:\s*(.*?)(?=\b3\s*:|$)", text, re.I)
    if first_match:
        first_text = first_match.group(1)
    if second_match:
        second_text = second_match.group(1)

    first = parse_language_option_list(first_text)
    second = parse_language_option_list(second_text)
    if not first and not second:
        all_options = parse_language_option_list(text)
    else:
        all_options = dedupe_language_options([*first, *second])
    return first, second, all_options


def parse_class_name_language_slots(
    value: Any,
) -> tuple[tuple[tuple[str, str], ...], tuple[tuple[str, str], ...]]:
    empty: tuple[tuple[str, str], ...] = tuple()
    text = safe_text(value)
    if "(" not in text:
        return empty, empty
    parts = [
        part.strip() for part in re.split(r"\s+-\s+", text, maxsplit=2) if part.strip()
    ]
    profile_part = parts[-1] if parts else text
    if "(" not in profile_part:
        return empty, empty
    lang_part = profile_part[profile_part.find("(") + 1 :].strip()
    while lang_part.endswith(")") and lang_part.count(")") > lang_part.count("("):
        lang_part = lang_part[:-1].strip()
    segments = [
        part.strip() for part in re.split(r"\s*-\s*", lang_part) if part.strip()
    ]
    parsed_segments: list[tuple[str, str]] = []
    for segment in segments:
        parsed = parse_language_option(segment)
        if parsed:
            parsed_segments.append(parsed)
    first = tuple(parsed_segments[:1])
    second = tuple(parsed_segments[1:])
    return first, second


def language_options_for_row(row: pd.Series) -> dict[str, tuple[tuple[str, str], ...]]:
    first = list(parse_language_option_list(row.get("PierwszyJezykObcy")))
    second = list(parse_language_option_list(row.get("DrugiJezykObcy")))
    legacy_first, legacy_second, legacy_all = parse_legacy_language_slots(
        row.get("JezykiObce")
    )
    name_first, name_second = parse_class_name_language_slots(row.get("OddzialNazwa"))
    icon_options = list(parse_language_option_list(row.get("JezykiObceIkonyOpis")))

    if not first:
        first.extend(legacy_first or name_first)
    if not second:
        second.extend(legacy_second or name_second)
    if not first:
        first.extend(option for option in icon_options if option[1] == "dwujęzyczny")

    all_options = dedupe_language_options(
        [*first, *second, *icon_options, *legacy_all, *name_first, *name_second]
    )
    return {
        "first": dedupe_language_options(first),
        "second": dedupe_language_options(second),
        "all": all_options,
    }


def language_display_values(options: tuple[tuple[str, str], ...], item: str) -> str:
    index = 0 if item == "language" else 1
    values = unique_preserving_order([option[index] for option in options])
    return "; ".join(values)


def language_pair_values(options: tuple[tuple[str, str], ...]) -> str:
    return "; ".join(
        f"{language}|{level}"
        for language, level in dedupe_language_options(list(options))
    )


def normalized_language_columns(row: pd.Series) -> dict[str, str]:
    options = language_options_for_row(row)
    return {
        "JezykiPierwszeNorm": language_display_values(options["first"], "language"),
        "JezykiDrugieNorm": language_display_values(options["second"], "language"),
        "JezykiWszystkieNorm": language_display_values(options["all"], "language"),
        "JezykiPierwszePoziomy": language_display_values(options["first"], "level"),
        "JezykiDrugiePoziomy": language_display_values(options["second"], "level"),
        "JezykiWszystkiePoziomy": language_display_values(options["all"], "level"),
        "JezykiPierwszeOpcje": language_pair_values(options["first"]),
        "JezykiDrugieOpcje": language_pair_values(options["second"]),
        "JezykiWszystkieOpcje": language_pair_values(options["all"]),
    }


def class_profile_text_from_name(value: Any) -> str:
    """Wyciąga część profilu z nazwy oddziału, bez typu i języków w nawiasie."""
    text = safe_text(value)
    if not text:
        return ""
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\[[^\]]*\]", " ", text)
    parts = [part.strip() for part in re.split(r"\s+-\s+", text) if part.strip()]
    if parts:
        text = parts[-1]
    text = re.sub(r"^\s*\d*\s*[A-Za-z]+\d*\s+", " ", text)
    return text


def class_subject_tokens(row: pd.Series) -> tuple[str, ...]:
    text = " ".join(
        safe_text(row.get(col))
        for col in [
            "PrzedmiotyRozszerzone",
            "Zawod",
            "DyscyplinaSportowa",
        ]
    )
    text = f"{text} {class_profile_text_from_name(row.get('OddzialNazwa'))}"
    return token_set_from_text(text, SUBJECT_MATCH_ALIASES)


def class_language_tokens(row: pd.Series) -> tuple[str, ...]:
    options = language_options_for_row(row)
    return tuple(
        sorted(
            {
                LANGUAGE_TOKEN_BY_DISPLAY[language]
                for language, _level in options["all"]
                if language in LANGUAGE_TOKEN_BY_DISPLAY
            }
        )
    )


def class_type_token(name: Any, explicit_type: Any = None) -> str:
    text = safe_text(name)
    match = re.search(r"\[([^\]]+)\]|\(([A-Z]{1,3}(?:/[io])?)\)", text)
    if match:
        return ascii_key(match.group(1) or match.group(2)).upper()
    explicit = safe_text(explicit_type)
    explicit_key = ascii_key(explicit)
    for key, value in CLASS_TYPE_ALIASES.items():
        if ascii_key(key) in explicit_key:
            return value
    return explicit_key.upper()


def jaccard_score(left: tuple[str, ...], right: tuple[str, ...]) -> float:
    left_set = set(left)
    right_set = set(right)
    if not left_set or not right_set:
        return 0.0
    return len(left_set & right_set) / len(left_set | right_set)


def text_similarity(left: Any, right: Any) -> float:
    return difflib.SequenceMatcher(None, ascii_key(left), ascii_key(right)).ratio()


def preview_text(value: Any, limit: int = 500) -> str:
    text = safe_text(value).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rsplit(" ", 1)[0].strip() + "..."


def normalize_address(value: Any) -> str:
    text = ascii_key(value)
    text = re.sub(r"\b(ul|al|aleja|pl|plac|warszawa)\b", " ", text)
    text = re.sub(r"\b\d{2}\s*\d{3}\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def pzo_search_url(school_type_ids: Any = "") -> str:
    first_type = safe_text(school_type_ids).split(",", 1)[0].strip()
    payload: dict[str, Any] = {
        "noRecrutation": False,
        "otherRecrutation": False,
        "freePlaces": False,
        "recrutationModule": True,
        "offerMap": {},
        "connective": {},
        "chosenOperatorMap": {},
    }
    if first_type:
        payload["schoolTypeId"] = int(first_type)
    query = quote(json.dumps(payload, ensure_ascii=False, sort_keys=True), safe="")
    return f"{PZO_BASE_URL}{PZO_PUBLIC_CONTEXT}/offer/search/results?q={query}"


def ensure_source_file(source: dict[str, Any]) -> Path:
    path = resolve_path(source["path"])
    if path.exists():
        return path
    url = source.get("source_url")
    if not url:
        raise FileNotFoundError(f"Brak lokalnego pliku i URL zrodla: {path}")
    logger.info("Pobieranie zrodla: %s", url)
    response = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
    response.raise_for_status()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(response.content)
    logger.info("Zapisano zrodlo do %s", path)
    return path


def threshold_sources(year_cfg: dict[str, Any]) -> list[dict[str, Any]]:
    threshold_cfg = year_cfg.get("thresholds")
    if not threshold_cfg:
        return []
    if "sources" not in threshold_cfg:
        source = threshold_cfg.copy()
        source.setdefault("priority", 1)
        return [source]

    inherited = {key: value for key, value in threshold_cfg.items() if key != "sources"}
    sources = []
    for index, source in enumerate(threshold_cfg["sources"], start=1):
        merged = inherited | source
        merged.setdefault("priority", index)
        sources.append(merged)
    return sources


def load_thresholds(year_cfg: dict[str, Any]) -> pd.DataFrame:
    sources = threshold_sources(year_cfg)
    if not sources:
        return pd.DataFrame(columns=THRESHOLD_COLUMNS)

    frames = []
    for source in sources:
        path = ensure_source_file(source)
        threshold_year = source.get("threshold_year", year_cfg.get("admission_year"))
        df_source = load_min_points(path, admission_year=threshold_year)
        if "admission_year" in df_source.columns:
            df_source = df_source.rename(columns={"admission_year": "threshold_year"})
        else:
            df_source["threshold_year"] = threshold_year
        df_source["threshold_kind"] = source.get(
            "threshold_kind", year_cfg.get("threshold_mode", "actual")
        )
        df_source["threshold_priority"] = int(source.get("priority", 1))
        df_source["threshold_label"] = source.get(
            "threshold_label", f"progi {threshold_year}"
        )
        df_source["threshold_source"] = source.get("source_url", str(path))
        df_source["SzkolaIdentyfikator"] = df_source["NazwaSzkoly"].apply(
            normalize_name
        )
        df_source["year"] = year_cfg["year"]
        df_source["admission_year"] = year_cfg.get("admission_year")
        df_source["school_year"] = year_cfg.get("school_year")
        frames.append(df_source)

    return pd.concat(frames, ignore_index=True, sort=False)


def best_thresholds_for_keys(
    df_thresholds: pd.DataFrame, keys: list[str]
) -> pd.DataFrame:
    if df_thresholds.empty:
        return df_thresholds.copy()
    required = keys + ["threshold_priority"]
    if not set(required).issubset(df_thresholds.columns):
        return df_thresholds.copy()
    sort_cols = keys + ["threshold_priority"]
    if "Prog_min_klasa" in df_thresholds.columns:
        sort_cols.append("Prog_min_klasa")
    return (
        df_thresholds.sort_values(sort_cols, na_position="last")
        .drop_duplicates(keys, keep="first")
        .copy()
    )


def school_threshold_summary(df_thresholds: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "SzkolaIdentyfikator",
        "Prog_min_szkola",
        "Prog_max_szkola",
        "Prog_szkola_threshold_year",
        "Prog_szkola_threshold_kind",
        "Prog_szkola_threshold_label",
    ]
    if df_thresholds.empty:
        return pd.DataFrame(columns=columns)

    school_source_cols = [
        "SzkolaIdentyfikator",
        "threshold_priority",
        "threshold_year",
        "threshold_kind",
        "threshold_label",
    ]
    best_school_sources = best_thresholds_for_keys(
        df_thresholds[school_source_cols].drop_duplicates(),
        ["SzkolaIdentyfikator"],
    )
    selected = df_thresholds.merge(
        best_school_sources[
            ["SzkolaIdentyfikator", "threshold_priority", "threshold_year"]
        ],
        how="inner",
        on=["SzkolaIdentyfikator", "threshold_priority", "threshold_year"],
    )
    return (
        selected.groupby("SzkolaIdentyfikator")
        .agg(
            Prog_min_szkola=("Prog_min_klasa", "min"),
            Prog_max_szkola=("Prog_min_klasa", "max"),
            Prog_szkola_threshold_year=("threshold_year", "first"),
            Prog_szkola_threshold_kind=("threshold_kind", "first"),
            Prog_szkola_threshold_label=("threshold_label", "first"),
        )
        .reset_index()
    )


def format_threshold_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    number = float(value)
    return str(int(number)) if number.is_integer() else f"{number:.2f}".rstrip("0")


def format_threshold_year(value: Any) -> str:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        number = float(value)
    except (TypeError, ValueError):
        return safe_text(value).strip()
    return str(int(number)) if number.is_integer() else str(number)


def format_threshold_range(min_value: Any, max_value: Any) -> str:
    min_text = format_threshold_value(min_value)
    max_text = format_threshold_value(max_value)
    if min_text == max_text:
        return min_text
    return f"{min_text}-{max_text}"


def format_ranking_value(value: Any, text_value: Any = None) -> str:
    if text_value is not None and pd.notna(text_value):
        text = str(text_value).strip()
        if text and text.lower() != "nan":
            text = text.rstrip("=")
            return text[:-2] if text.endswith(".0") else text
    if pd.isna(value):
        return ""
    number = float(value)
    return str(int(number)) if number.is_integer() else f"{number:.2f}".rstrip("0")


def historical_school_thresholds(df_thresholds: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "SzkolaIdentyfikator",
        "Progi_historyczne_szkola",
        "Progi_historyczne_lata",
    ]
    if df_thresholds.empty:
        return pd.DataFrame(columns=columns)

    ranges = (
        df_thresholds.groupby(["SzkolaIdentyfikator", "threshold_year"])
        .agg(
            Prog_min=("Prog_min_klasa", "min"),
            Prog_max=("Prog_min_klasa", "max"),
        )
        .reset_index()
        .sort_values(["SzkolaIdentyfikator", "threshold_year"], ascending=[True, False])
    )

    rows = []
    for school_id, group in ranges.groupby("SzkolaIdentyfikator", sort=False):
        parts = [
            f"{int(row.threshold_year)}: "
            f"{format_threshold_range(row.Prog_min, row.Prog_max)}"
            for row in group.itertuples(index=False)
        ]
        rows.append(
            {
                "SzkolaIdentyfikator": school_id,
                "Progi_historyczne_szkola": "; ".join(parts),
                "Progi_historyczne_lata": "/".join(
                    str(int(year)) for year in group["threshold_year"].tolist()
                ),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def school_ranking_summary(df_rankings: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "SzkolaIdentyfikator",
        "RankingPozNajnowszy",
        "RankingPozTekstNajnowszy",
        "RankingRok",
        "Ranking_historyczny_szkola",
        "Ranking_lata",
    ]
    if df_rankings.empty or "SzkolaIdentyfikator" not in df_rankings.columns:
        return pd.DataFrame(columns=columns)

    rankings = df_rankings.copy()
    if "RankingPozTekst" not in rankings.columns:
        rankings["RankingPozTekst"] = rankings["RankingPoz"].astype(str)
    rankings["RankingPoz"] = pd.to_numeric(rankings["RankingPoz"], errors="coerce")
    rankings["year"] = pd.to_numeric(rankings["year"], errors="coerce")
    rankings = rankings.dropna(subset=["SzkolaIdentyfikator", "year", "RankingPoz"])
    if rankings.empty:
        return pd.DataFrame(columns=columns)

    rankings = (
        rankings.sort_values(
            ["SzkolaIdentyfikator", "year", "RankingPoz"],
            ascending=[True, False, True],
        )
        .drop_duplicates(["SzkolaIdentyfikator", "year"], keep="first")
        .sort_values(["SzkolaIdentyfikator", "year"], ascending=[True, False])
    )

    rows = []
    for school_id, group in rankings.groupby("SzkolaIdentyfikator", sort=False):
        latest = group.iloc[0]
        parts = [
            f"{int(row.year)}: "
            f"{format_ranking_value(row.RankingPoz, row.RankingPozTekst)}"
            for row in group.itertuples(index=False)
        ]
        rows.append(
            {
                "SzkolaIdentyfikator": school_id,
                "RankingPozNajnowszy": latest["RankingPoz"],
                "RankingPozTekstNajnowszy": format_ranking_value(
                    latest["RankingPoz"], latest["RankingPozTekst"]
                ),
                "RankingRok": int(latest["year"]),
                "Ranking_historyczny_szkola": "; ".join(parts),
                "Ranking_lata": "/".join(
                    str(int(year)) for year in group["year"].tolist()
                ),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def apply_latest_rankings(sheets: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    ranking_summary = school_ranking_summary(sheets.get("rankings", pd.DataFrame()))
    if ranking_summary.empty:
        return sheets

    updated = sheets.copy()
    for sheet_name in ["schools", "classes"]:
        df = updated.get(sheet_name, pd.DataFrame())
        if df.empty or "SzkolaIdentyfikator" not in df.columns:
            continue
        df = df.copy()
        rename_columns = {}
        if "RankingPoz" in df.columns:
            rename_columns["RankingPoz"] = "RankingPozRokuDanych"
        if "RankingPozTekst" in df.columns:
            rename_columns["RankingPozTekst"] = "RankingPozTekstRokuDanych"
        df = df.rename(columns=rename_columns)
        df = df.drop(
            columns=[
                "RankingPozNajnowszy",
                "RankingPozTekstNajnowszy",
                "RankingRok",
                "Ranking_historyczny_szkola",
                "Ranking_lata",
            ],
            errors="ignore",
        )
        df = df.merge(ranking_summary, how="left", on="SzkolaIdentyfikator")
        df["RankingPoz"] = df["RankingPozNajnowszy"]
        df["RankingPozTekst"] = df["RankingPozTekstNajnowszy"]
        updated[sheet_name] = df
    return updated


def read_app_workbook_sheets(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        return {}
    excel = pd.ExcelFile(path)
    return {
        sheet_name: pd.read_excel(excel, sheet_name=sheet_name)
        for sheet_name in excel.sheet_names
    }


def restore_year_ranking_columns(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    if sheet_name not in {"schools", "classes"} or df.empty:
        return df

    df = df.copy()
    if "RankingPozRokuDanych" in df.columns:
        df["RankingPoz"] = df["RankingPozRokuDanych"]
    if "RankingPozTekstRokuDanych" in df.columns:
        df["RankingPozTekst"] = df["RankingPozTekstRokuDanych"]
    return df.drop(
        columns=[
            "RankingPozRokuDanych",
            "RankingPozTekstRokuDanych",
            "RankingPozNajnowszy",
            "RankingPozTekstNajnowszy",
            "RankingRok",
            "Ranking_historyczny_szkola",
            "Ranking_lata",
        ],
        errors="ignore",
    )


def merge_existing_year_sheets(
    existing_sheets: dict[str, pd.DataFrame],
    new_sheets: dict[str, pd.DataFrame],
    replace_years: set[int],
) -> dict[str, pd.DataFrame]:
    merged_sheets = {}
    sheet_names = list(new_sheets)
    sheet_names.extend(name for name in existing_sheets if name not in new_sheets)
    for sheet_name in sheet_names:
        existing = restore_year_ranking_columns(
            sheet_name, existing_sheets.get(sheet_name, pd.DataFrame())
        )
        new = new_sheets.get(sheet_name, pd.DataFrame())

        if existing.empty:
            merged_sheets[sheet_name] = new
            continue
        if new.empty and "year" not in existing.columns:
            merged_sheets[sheet_name] = existing
            continue
        if "year" in existing.columns:
            existing_year = pd.to_numeric(existing["year"], errors="coerce")
            existing = existing[~existing_year.isin(replace_years)].copy()
        if new.empty:
            merged_sheets[sheet_name] = existing
        elif existing.empty:
            merged_sheets[sheet_name] = new
        else:
            merged_sheets[sheet_name] = pd.concat(
                [existing, new], ignore_index=True, sort=False
            )
    return merged_sheets


def threshold_meta(year_cfg: dict[str, Any]) -> dict[str, Any]:
    sources = threshold_sources(year_cfg)
    years = [
        str(source.get("threshold_year"))
        for source in sources
        if source.get("threshold_year") is not None
    ]
    source_refs = [source.get("source_url") or source.get("path") for source in sources]
    return {
        "threshold_mode": year_cfg.get("threshold_mode"),
        "threshold_label": year_cfg.get("threshold_label"),
        "threshold_years": "/".join(years),
        "threshold_source": " | ".join(str(ref) for ref in source_refs if ref),
    }


def add_year_metadata(
    df: pd.DataFrame, year_cfg: dict[str, Any], threshold_info: dict[str, Any]
) -> None:
    df["year"] = year_cfg["year"]
    df["admission_year"] = year_cfg.get("admission_year")
    df["school_year"] = year_cfg.get("school_year")
    df["data_status"] = year_cfg.get("data_status")
    df["status_label"] = year_cfg.get("status_label")
    df["threshold_mode"] = threshold_info["threshold_mode"]
    if "threshold_label" not in df.columns:
        df["threshold_label"] = threshold_info["threshold_label"]
    else:
        empty_label = df["threshold_label"].apply(
            lambda value: not safe_text(value).strip()
        )
        df.loc[empty_label, "threshold_label"] = threshold_info["threshold_label"]
    if "threshold_years" not in df.columns:
        df["threshold_years"] = threshold_info["threshold_years"]


def load_ranking(year_cfg: dict[str, Any]) -> pd.DataFrame:
    ranking_cfg = year_cfg.get("ranking")
    if not ranking_cfg:
        return pd.DataFrame(columns=["RankingPoz", "NazwaSzkoly", "Dzielnica"])

    cache_path = (
        resolve_path(ranking_cfg["cache_path"])
        if ranking_cfg.get("cache_path")
        else None
    )
    if cache_path and cache_path.exists():
        df = pd.read_excel(cache_path)
    else:
        source_type = ranking_cfg["type"]
        path = ensure_source_file(ranking_cfg)
        if source_type == "perspektywy_pdf":
            df = parse_ranking_perspektywy_pdf(path)
        elif source_type == "perspektywy_html":
            df = parse_ranking_perspektywy_html(path, year=year_cfg["year"])
        else:
            raise ValueError(f"Nieznany typ rankingu: {source_type}")
        if cache_path:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(cache_path, index=False)

    if "RankingPozTekst" not in df.columns and "RankingPoz" in df.columns:
        df["RankingPozTekst"] = df["RankingPoz"].astype(str)
    df["RankingPoz"] = pd.to_numeric(df["RankingPoz"], errors="coerce")
    df["SzkolaIdentyfikator"] = df["NazwaSzkoly"].apply(normalize_name)
    df["year"] = year_cfg["year"]
    df["school_year"] = year_cfg.get("school_year")
    return df


def load_location_cache() -> pd.DataFrame:
    candidates = [
        resolve_path("results/app/licea_warszawa.xlsx"),
        LEGACY_APP_FILE,
        RESULTS_DIR / "LO_Warszawa_2025_Warszawa_Metro_Wilanowska.xlsx",
    ]
    frames = []
    for path in candidates:
        if not path.exists():
            continue
        try:
            excel = pd.ExcelFile(path)
            sheet = "schools" if "schools" in excel.sheet_names else "szkoly"
            df = pd.read_excel(excel, sheet_name=sheet)
        except Exception as exc:
            logger.warning(
                "Nie udało się wczytać cache lokalizacji z %s: %s", path, exc
            )
            continue
        cols = [
            col
            for col in [
                "source_school_id",
                "SzkolaIdentyfikator",
                "NazwaSzkoly",
                "AdresSzkoly",
                "TypSzkoly",
                "year",
                "CzasDojazdu",
                "SzkolaLat",
                "SzkolaLon",
                "url",
            ]
            if col in df.columns
        ]
        if cols:
            frames.append(df[cols].copy())
    if not frames:
        return pd.DataFrame()
    cache = pd.concat(frames, ignore_index=True)
    if "source_school_id" in cache.columns and cache["source_school_id"].notna().any():
        cache = cache.dropna(subset=["source_school_id"]).drop_duplicates(
            subset=["source_school_id"], keep="first"
        )
    else:
        cache = cache.dropna(subset=["SzkolaIdentyfikator"]).drop_duplicates(
            subset=["SzkolaIdentyfikator"], keep="first"
        )
    return cache


def load_vulcan_offer(year_cfg: dict[str, Any]) -> pd.DataFrame:
    offer_cfg = year_cfg["offer"]
    path = resolve_path(offer_cfg["path"])
    if path.exists():
        return pd.read_excel(path)

    from scripts.data_processing.get_data_vulcan_async import download_all_async

    start_id = offer_cfg.get("start_id", 1)
    end_id = offer_cfg.get("end_id", 400)
    df = asyncio.run(download_all_async(start_id, end_id, verbose=True))
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)
    return df


def prepare_vulcan_offer(df_vulcan: pd.DataFrame, cfg: dict[str, Any]) -> pd.DataFrame:
    df_vulcan = df_vulcan.copy()
    df_vulcan["TypSzkoly"] = df_vulcan["NazwaSzkoly"].apply(get_school_type)

    filtr_miasto = cfg.get("filtr_miasto")
    if filtr_miasto:
        df_vulcan = df_vulcan[
            df_vulcan["AdresSzkoly"].str.contains(filtr_miasto, na=False, case=False)
        ]

    filtr_typ = cfg.get("filtr_typ_szkola")
    if filtr_typ:
        typy = [filtr_typ] if isinstance(filtr_typ, str) else filtr_typ
        df_vulcan = df_vulcan[df_vulcan["TypSzkoly"].isin(typy)]

    df_vulcan["SzkolaIdentyfikator"] = df_vulcan["NazwaSzkoly"].apply(normalize_name)
    df_vulcan["source_school_id"] = "vulcan:" + df_vulcan["IdSzkoly"].astype(str)
    df_vulcan["Kod"] = df_vulcan["AdresSzkoly"].str.extract(r"(\d{2}-\d{3})")
    if KODY_FILE.exists():
        df_pc = pd.read_csv(KODY_FILE, dtype=str)
        df_vulcan = df_vulcan.merge(df_pc, how="left", on="Kod")
    else:
        df_vulcan["Dzielnica"] = None
    return df_vulcan


def attach_location_data(
    df_schools: pd.DataFrame,
    cfg: dict[str, Any],
    location_cache: pd.DataFrame | None = None,
) -> pd.DataFrame:
    df_schools = df_schools.copy()
    api_key = os.getenv("GOOGLE_MAPS_API_KEY")
    should_fetch = bool(cfg.get("pobierz_nowe_czasy", True) and api_key)

    if should_fetch:
        try:
            import googlemaps
        except ImportError:
            logger.warning("Brak pakietu googlemaps; używam cache lokalizacji.")
        else:
            try:
                df_schools["PelenAdres"] = (
                    df_schools["NazwaSzkoly"].str.strip()
                    + ", "
                    + df_schools["AdresSzkoly"].str.strip()
                )
                addresses = df_schools["PelenAdres"].dropna().unique().tolist()
                client = googlemaps.Client(key=api_key)
                departure_timestamp = get_next_weekday_time(
                    cfg.get("departure_hour", 7), cfg.get("departure_minute", 30)
                )
                travel_times = {}
                batch_size = cfg.get("googlemaps_batch_size", 25)
                for i in range(0, len(addresses), batch_size):
                    batch = addresses[i : i + batch_size]
                    travel_times.update(
                        get_travel_times_batch(
                            gmaps=client,
                            origin_address=cfg["adres_domowy"],
                            destination_addresses=batch,
                            mode="transit",
                            departure_time=departure_timestamp,
                        )
                    )
                    time.sleep(0.25)
                coordinates = get_coordinates_for_addresses_batch(client, addresses)
                df_schools["CzasDojazdu"] = df_schools["PelenAdres"].map(travel_times)
                df_schools["SzkolaLat"] = df_schools["PelenAdres"].map(
                    lambda addr: coordinates.get(addr, (None, None))[0]
                )
                df_schools["SzkolaLon"] = df_schools["PelenAdres"].map(
                    lambda addr: coordinates.get(addr, (None, None))[1]
                )
                df_schools.drop(columns=["PelenAdres"], inplace=True)
                df_schools[
                    [
                        "SzkolaIdentyfikator",
                        "AdresSzkoly",
                        "CzasDojazdu",
                        "SzkolaLat",
                        "SzkolaLon",
                    ]
                ].drop_duplicates().to_excel(CZASY_DOJAZDU_FILE, index=False)
                return df_schools
            except Exception as exc:
                logger.warning(
                    "Błąd przy pobieraniu danych Google Maps (%s); używam cache lokalizacji.",
                    exc,
                )
                df_schools.drop(columns=["PelenAdres"], errors="ignore", inplace=True)

    if location_cache is None or location_cache.empty:
        if CZASY_DOJAZDU_FILE.exists():
            location_cache = pd.read_excel(CZASY_DOJAZDU_FILE)
        else:
            location_cache = pd.DataFrame()

    if not location_cache.empty and "SzkolaIdentyfikator" in location_cache.columns:
        merge_key = "SzkolaIdentyfikator"
        if (
            "source_school_id" in df_schools.columns
            and "source_school_id" in location_cache.columns
            and location_cache["source_school_id"].notna().any()
        ):
            merge_key = "source_school_id"
        cols = [
            col
            for col in [
                merge_key,
                "CzasDojazdu",
                "SzkolaLat",
                "SzkolaLon",
            ]
            if col in location_cache.columns
        ]
        df_schools = df_schools.drop(
            columns=[col for col in cols if col != merge_key],
            errors="ignore",
        )
        df_schools = df_schools.merge(
            location_cache[cols].drop_duplicates(merge_key),
            on=merge_key,
            how="left",
        )

    for col in ["CzasDojazdu", "SzkolaLat", "SzkolaLon"]:
        if col not in df_schools.columns:
            df_schools[col] = None
    return df_schools


def add_common_class_columns(df_classes: pd.DataFrame) -> pd.DataFrame:
    df_classes = df_classes.copy()
    df_classes["Profil"] = (
        df_classes["OddzialNazwa"]
        .astype(str)
        .str.extract(r"\[[^\]]+\]\s*([^\(]+)")[0]
        .str.strip()
    )
    parsed_class_type = df_classes["OddzialNazwa"].apply(extract_class_type)
    if "TypOddzialu" in df_classes.columns:
        existing_class_type = df_classes["TypOddzialu"].replace("", pd.NA)
        df_classes["TypOddzialu"] = existing_class_type.combine_first(parsed_class_type)
    else:
        df_classes["TypOddzialu"] = parsed_class_type
    language_columns = df_classes.apply(
        normalized_language_columns, axis=1, result_type="expand"
    )
    for column in language_columns.columns:
        df_classes[column] = language_columns[column].fillna("")
    if "JezykiObce" in df_classes.columns:
        df_classes["JezykiObce"] = (
            df_classes["JezykiObce"]
            .fillna("")
            .astype(str)
            .str.replace("Pierwszy", "1", regex=False)
            .str.replace("Drugi", "2", regex=False)
            .str.replace("Trzeci", "3", regex=False)
            .str.replace("język", "", regex=False)
        )
    else:
        df_classes["JezykiObce"] = ""

    if "PrzedmiotyRozszerzone" not in df_classes.columns:
        df_classes["PrzedmiotyRozszerzone"] = ""
    for subject in ALL_SUBJECTS:
        pattern = rf"(?i)\b{re.escape(subject)}\b"
        df_classes[subject] = (
            df_classes["PrzedmiotyRozszerzone"]
            .fillna("")
            .astype(str)
            .str.contains(pattern, na=False)
            .astype(int)
        )
    return df_classes


def load_pzo_offer_tables(year_cfg: dict[str, Any]) -> dict[str, pd.DataFrame]:
    offer_cfg = year_cfg["offer"]
    path = resolve_path(offer_cfg["path"])
    if path.is_dir():
        return build_pzo_tables(load_pzo_snapshot_files(path))
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and path.exists():
        excel = pd.ExcelFile(path)
        return {
            sheet: pd.read_excel(excel, sheet_name=sheet) for sheet in excel.sheet_names
        }
    if offer_cfg.get("auto_download", True):
        raw_dir = path if not path.suffix else path.with_suffix("")
        logger.info(
            "Brak lokalnego snapshotu PZO w %s; pobieram publiczny snapshot.",
            raw_dir,
        )
        client = PzoOmikronClient(
            base_url=offer_cfg.get("base_url", PZO_BASE_URL),
            public_context=offer_cfg.get("public_context", PZO_PUBLIC_CONTEXT),
            timeout=int(offer_cfg.get("timeout", 60)),
        )
        snapshot = fetch_offer_snapshot(
            client=client,
            year=int(year_cfg["year"]),
            school_year=year_cfg.get("school_year", ""),
            school_type_ids=offer_cfg.get("school_type_ids"),
            limit_schools=offer_cfg.get("limit_schools"),
            delay=float(offer_cfg.get("delay", 0.0)),
        )
        write_snapshot_files(snapshot, raw_dir)
        return build_pzo_tables(snapshot)
    raise FileNotFoundError(
        "Brak lokalnego snapshotu PZO. Uruchom najpierw "
        "scripts/data_processing/get_data_pzo_omikron.py i sprawdź path w data_sources.yml: "
        f"{path}"
    )


def reference_schools_from_cache(location_cache: pd.DataFrame) -> pd.DataFrame:
    if location_cache.empty or "SzkolaIdentyfikator" not in location_cache.columns:
        return pd.DataFrame()
    refs = location_cache.copy()
    if "year" in refs.columns:
        years = pd.to_numeric(refs["year"], errors="coerce")
        refs = refs[years.eq(2025)]
    required = ["SzkolaIdentyfikator", "NazwaSzkoly", "AdresSzkoly", "TypSzkoly"]
    missing = [col for col in required if col not in refs.columns]
    if missing:
        return pd.DataFrame()
    refs = refs.dropna(subset=["SzkolaIdentyfikator", "NazwaSzkoly"])
    return refs[required].drop_duplicates("SzkolaIdentyfikator")


def attach_stable_school_ids(
    df_schools: pd.DataFrame,
    df_classes: pd.DataFrame,
    reference_schools: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    schools = df_schools.copy()
    classes = df_classes.copy()
    if reference_schools.empty:
        schools["SzkolaIdentyfikator"] = schools["NazwaSzkoly"].apply(normalize_name)
        schools["PzoSchoolMatchStatus"] = "fallback_name"
        schools["PzoSchoolMatchScore"] = pd.NA
    else:
        refs = reference_schools.copy()
        refs["addr_key"] = refs["AdresSzkoly"].apply(normalize_address)
        refs["name_key"] = refs["NazwaSzkoly"].apply(ascii_key)
        refs["type_key"] = refs["TypSzkoly"].apply(ascii_key)

        mapped_rows = []
        for row in schools.itertuples(index=False):
            address_key = normalize_address(getattr(row, "AdresSzkoly", ""))
            type_key = ascii_key(getattr(row, "TypSzkoly", ""))
            name_key = ascii_key(getattr(row, "NazwaSzkoly", ""))
            candidates = refs.copy()
            if type_key and "type_key" in candidates:
                same_type = candidates[candidates["type_key"].eq(type_key)]
                if not same_type.empty:
                    candidates = same_type

            exact_address = candidates[candidates["addr_key"].eq(address_key)]
            if not exact_address.empty:
                scored = exact_address.copy()
                scored["score"] = scored["name_key"].apply(
                    lambda value: difflib.SequenceMatcher(None, name_key, value).ratio()
                )
                best = scored.sort_values("score", ascending=False).iloc[0]
                mapped_rows.append(
                    {
                        "source_school_id": row.source_school_id,
                        "SzkolaIdentyfikator": best["SzkolaIdentyfikator"],
                        "PzoSchoolMatchStatus": "same_address",
                        "PzoSchoolMatchScore": float(best["score"]),
                    }
                )
                continue

            scored = candidates.copy()
            scored["name_score"] = scored["name_key"].apply(
                lambda value: difflib.SequenceMatcher(None, name_key, value).ratio()
            )
            scored["addr_score"] = scored["addr_key"].apply(
                lambda value: difflib.SequenceMatcher(None, address_key, value).ratio()
            )
            scored["score"] = scored["name_score"] * 0.65 + scored["addr_score"] * 0.35
            best = scored.sort_values("score", ascending=False).iloc[0]
            if float(best["score"]) >= 0.62:
                mapped_rows.append(
                    {
                        "source_school_id": row.source_school_id,
                        "SzkolaIdentyfikator": best["SzkolaIdentyfikator"],
                        "PzoSchoolMatchStatus": "name_address_similarity",
                        "PzoSchoolMatchScore": float(best["score"]),
                    }
                )
            else:
                mapped_rows.append(
                    {
                        "source_school_id": row.source_school_id,
                        "SzkolaIdentyfikator": normalize_name(row.NazwaSzkoly),
                        "PzoSchoolMatchStatus": "fallback_name",
                        "PzoSchoolMatchScore": float(best["score"]),
                    }
                )

        mapping = pd.DataFrame(mapped_rows)
        schools = schools.drop(columns=["SzkolaIdentyfikator"], errors="ignore").merge(
            mapping, on="source_school_id", how="left"
        )

    classes = classes.drop(columns=["SzkolaIdentyfikator"], errors="ignore").merge(
        schools[
            [
                "source_school_id",
                "SzkolaIdentyfikator",
                "PzoSchoolMatchStatus",
                "PzoSchoolMatchScore",
            ]
        ],
        on="source_school_id",
        how="left",
    )
    return schools, classes


def parse_pointed_subjects(text: str) -> dict[str, str]:
    result: dict[str, str] = {}
    if not text:
        return result
    ordinal_pattern = "|".join(ORDINAL_TO_COLUMN)
    pattern = re.compile(
        rf"({ordinal_pattern})\s+punktowany\s+przedmiot\s*:?\s*(.+?)"
        rf"(?=\s+(?:{ordinal_pattern})\s+punktowany\s+przedmiot|$)",
        flags=re.IGNORECASE,
    )
    for match in pattern.finditer(text):
        ordinal = ascii_key(match.group(1))
        column = ORDINAL_TO_COLUMN.get(ordinal)
        if not column:
            continue
        value = safe_text(match.group(2))
        value = re.sub(r"\s+", " ", value.replace(":", " ")).strip(" ;,.")
        if value:
            result[column] = value
    return result


def summarize_criteria(criteria: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "source_class_id",
        "Punktowany1",
        "Punktowany2",
        "Punktowany3",
        "Punktowany4",
        "PrzedmiotyPunktowane",
        "KryteriaPunktowane",
    ]
    if criteria.empty or "source_class_id" not in criteria.columns:
        return pd.DataFrame(columns=columns)
    rows: list[dict[str, Any]] = []
    for class_id, group in criteria.groupby("source_class_id", sort=False):
        texts: list[str] = []
        for col in ["group_header_text", "display_value_text"]:
            if col in group.columns:
                texts.extend(
                    safe_text(value)
                    for value in group[col].dropna().tolist()
                    if safe_text(value)
                )
        unique_texts = list(dict.fromkeys(texts))
        combined = " ".join(unique_texts)
        pointed = parse_pointed_subjects(combined)
        subjects = [
            pointed.get(col, "")
            for col in ["Punktowany1", "Punktowany2", "Punktowany3", "Punktowany4"]
            if pointed.get(col)
        ]
        rows.append(
            {
                "source_class_id": class_id,
                "Punktowany1": pointed.get("Punktowany1", ""),
                "Punktowany2": pointed.get("Punktowany2", ""),
                "Punktowany3": pointed.get("Punktowany3", ""),
                "Punktowany4": pointed.get("Punktowany4", ""),
                "PrzedmiotyPunktowane": ", ".join(subjects),
                "KryteriaPunktowane": "; ".join(unique_texts),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def prepare_threshold_features(df_thresholds: pd.DataFrame) -> pd.DataFrame:
    thresholds = df_thresholds.copy()
    if thresholds.empty:
        return thresholds
    if "SymbolOddzialu" not in thresholds.columns:
        thresholds["SymbolOddzialu"] = ""
    thresholds["match_code"] = thresholds.apply(
        lambda row: extract_class_code(
            row.get("OddzialNazwa"), row.get("SymbolOddzialu")
        ),
        axis=1,
    )
    thresholds["match_type"] = thresholds.apply(
        lambda row: class_type_token(row.get("OddzialNazwa")), axis=1
    )
    thresholds["match_subjects"] = thresholds.apply(class_subject_tokens, axis=1)
    thresholds["match_languages"] = thresholds.apply(class_language_tokens, axis=1)
    thresholds["match_name_key"] = thresholds["OddzialNazwa"].apply(ascii_key)
    return thresholds


def prepare_current_class_features(df_classes: pd.DataFrame) -> pd.DataFrame:
    classes = df_classes.copy()
    if classes.empty:
        return classes
    if "OddzialKod" not in classes.columns:
        classes["OddzialKod"] = ""
    classes["match_code"] = classes.apply(
        lambda row: extract_class_code(row.get("OddzialNazwa"), row.get("OddzialKod")),
        axis=1,
    )
    classes["match_type"] = classes.apply(
        lambda row: class_type_token(row.get("OddzialNazwa"), row.get("TypOddzialu")),
        axis=1,
    )
    classes["match_subjects"] = classes.apply(class_subject_tokens, axis=1)
    classes["match_languages"] = classes.apply(class_language_tokens, axis=1)
    classes["match_name_key"] = classes["OddzialNazwa"].apply(ascii_key)
    return classes


def threshold_match_status(
    score: float,
    gap: float,
    code_score: float,
    profile_score: float,
    language_score: float,
    type_score: float,
    threshold_value: Any,
) -> str:
    if pd.isna(threshold_value):
        return "candidate_only"
    profile_is_strong = profile_score >= 0.99
    profile_is_close = profile_score >= 0.66
    language_is_compatible = language_score >= 0.66
    language_has_overlap = language_score > 0
    type_is_compatible = type_score >= 1.0
    if (
        score >= 0.86
        and code_score >= 0.9
        and profile_is_strong
        and language_is_compatible
        and type_is_compatible
    ):
        return "trusted"
    if (
        score >= 0.72
        and profile_is_strong
        and language_is_compatible
        and type_is_compatible
        and gap >= 0.03
    ):
        return "approximate"
    if (
        score >= 0.74
        and profile_is_strong
        and language_has_overlap
        and type_is_compatible
        and gap >= 0.03
    ):
        return "approximate"
    if score >= 0.72 and code_score >= 0.75 and profile_is_close:
        return "approximate"
    return "candidate_only"


def match_reference_thresholds(
    df_classes: pd.DataFrame,
    df_thresholds: pd.DataFrame,
    max_candidates_per_class: int = 5,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    match_columns = [
        "source_school_id",
        "source_class_id",
        "SzkolaIdentyfikator",
        "OddzialNazwa",
        "threshold_year",
        "threshold_label",
        "threshold_kind",
        "threshold_priority",
        "OldOddzialNazwa",
        "OldSymbolOddzialu",
        "Prog_min_klasa",
        "match_score",
        "match_gap",
        "match_status",
        "match_method",
        "used_for_scoring",
        "candidate_rank",
    ]
    if df_classes.empty or df_thresholds.empty:
        return pd.DataFrame(columns=match_columns), pd.DataFrame()

    thresholds = df_thresholds.copy()
    thresholds["threshold_priority"] = pd.to_numeric(
        thresholds["threshold_priority"], errors="coerce"
    ).fillna(999)
    active_priority = thresholds.groupby("SzkolaIdentyfikator")[
        "threshold_priority"
    ].transform("min")
    thresholds = thresholds[thresholds["threshold_priority"].eq(active_priority)]
    thresholds = prepare_threshold_features(thresholds)
    classes = prepare_current_class_features(df_classes)

    rows = []
    for _, class_row in classes.iterrows():
        school_id = class_row.get("SzkolaIdentyfikator")
        if not safe_text(school_id):
            continue
        pool = thresholds[thresholds["SzkolaIdentyfikator"].eq(school_id)]
        scored = []
        for _, threshold_row in pool.iterrows():
            code_score, code_method = class_code_similarity(
                class_row.get("match_code"), threshold_row.get("match_code")
            )
            profile_score = jaccard_score(
                class_row.get("match_subjects", ()),
                threshold_row.get("match_subjects", ()),
            )
            language_score = jaccard_score(
                class_row.get("match_languages", ()),
                threshold_row.get("match_languages", ()),
            )
            type_score = (
                1.0
                if safe_text(class_row.get("match_type"))
                and class_row.get("match_type") == threshold_row.get("match_type")
                else 0.0
            )
            name_score = difflib.SequenceMatcher(
                None,
                safe_text(class_row.get("match_name_key")),
                safe_text(threshold_row.get("match_name_key")),
            ).ratio()
            score = (
                0.20 * code_score
                + 0.50 * profile_score
                + 0.15 * language_score
                + 0.10 * type_score
                + 0.05 * name_score
            )
            methods = [
                method
                for method, active in [
                    (code_method, bool(code_method)),
                    ("profile_exact", profile_score >= 0.99),
                    ("profile_partial", 0.66 <= profile_score < 0.99),
                    ("language_exact", language_score >= 0.99),
                    ("language_partial", 0 < language_score < 0.99),
                    ("type", type_score >= 1.0),
                ]
                if active
            ]
            scored.append(
                {
                    "source_school_id": class_row.get("source_school_id"),
                    "source_class_id": class_row.get("source_class_id"),
                    "SzkolaIdentyfikator": school_id,
                    "OddzialNazwa": class_row.get("OddzialNazwa"),
                    "threshold_year": threshold_row.get("threshold_year"),
                    "threshold_label": threshold_row.get("threshold_label"),
                    "threshold_kind": threshold_row.get("threshold_kind"),
                    "threshold_priority": threshold_row.get("threshold_priority"),
                    "OldOddzialNazwa": threshold_row.get("OddzialNazwa"),
                    "OldSymbolOddzialu": threshold_row.get("SymbolOddzialu"),
                    "Prog_min_klasa": threshold_row.get("Prog_min_klasa"),
                    "match_score": round(score, 4),
                    "code_score": code_score,
                    "profile_score": profile_score,
                    "language_score": language_score,
                    "type_score": type_score,
                    "match_method": ";".join(methods),
                }
            )
        if not scored:
            continue
        scored = sorted(
            scored,
            key=lambda item: (
                item["match_score"],
                item["profile_score"],
                item["language_score"],
                item["code_score"],
                -int(item.get("threshold_priority") or 999),
            ),
            reverse=True,
        )
        runner_score = scored[1]["match_score"] if len(scored) > 1 else 0.0
        for rank, item in enumerate(scored[:max_candidates_per_class], start=1):
            gap = float(item["match_score"]) - float(runner_score) if rank == 1 else 0.0
            item["match_gap"] = round(gap, 4)
            item["match_status"] = threshold_match_status(
                float(item["match_score"]),
                gap,
                float(item.pop("code_score")),
                float(item.pop("profile_score")),
                float(item.pop("language_score")),
                float(item.pop("type_score")),
                item.get("Prog_min_klasa"),
            )
            item["candidate_rank"] = rank
            item["used_for_scoring"] = rank == 1 and item["match_status"] in {
                "trusted",
                "approximate",
            }
            rows.append(item)

    matches = pd.DataFrame(rows, columns=match_columns)
    selected = matches[matches["used_for_scoring"].eq(True)].copy()
    return matches, selected


def apply_threshold_matches(
    df_classes: pd.DataFrame, threshold_matches: pd.DataFrame
) -> pd.DataFrame:
    classes = df_classes.copy()
    classes["Prog_min_klasa"] = pd.NA
    classes["ProgMatchStatus"] = "school_only"
    classes["ProgMatchScore"] = pd.NA
    classes["ProgMatchMethod"] = ""
    classes["ProgMatchOldClass"] = ""
    classes["ProgMatchLabel"] = ""
    classes["ProgCandidatesCount"] = 0
    classes["ProgCandidatesSummary"] = ""

    if threshold_matches.empty:
        return classes

    counts = (
        threshold_matches.groupby("source_class_id")
        .size()
        .rename("ProgCandidatesCount")
        .reset_index()
    )
    summary_rows = []
    for class_id, group in threshold_matches.sort_values(
        ["source_class_id", "candidate_rank"]
    ).groupby("source_class_id", sort=False):
        summary = "; ".join(
            f"{int(row.threshold_year)} {row.OldOddzialNazwa}: "
            f"{format_threshold_value(row.Prog_min_klasa)}"
            for row in group.head(3).itertuples(index=False)
            if pd.notna(row.Prog_min_klasa)
        )
        summary_rows.append(
            {"source_class_id": class_id, "ProgCandidatesSummary": summary}
        )
    candidate_summary = pd.DataFrame(summary_rows)
    classes = classes.merge(
        counts, on="source_class_id", how="left", suffixes=("", "_m")
    )
    classes["ProgCandidatesCount"] = classes["ProgCandidatesCount_m"].combine_first(
        classes["ProgCandidatesCount"]
    )
    classes = classes.drop(columns=["ProgCandidatesCount_m"], errors="ignore")
    classes = classes.merge(
        candidate_summary, on="source_class_id", how="left", suffixes=("", "_m")
    )
    classes["ProgCandidatesSummary"] = classes["ProgCandidatesSummary_m"].combine_first(
        classes["ProgCandidatesSummary"]
    )
    classes = classes.drop(columns=["ProgCandidatesSummary_m"], errors="ignore")

    selected = threshold_matches[threshold_matches["used_for_scoring"].eq(True)].copy()
    if selected.empty:
        return classes
    selected = selected[
        [
            "source_class_id",
            "Prog_min_klasa",
            "threshold_year",
            "threshold_kind",
            "threshold_label",
            "match_status",
            "match_score",
            "match_method",
            "OldOddzialNazwa",
        ]
    ].rename(
        columns={
            "match_status": "ProgMatchStatus",
            "match_score": "ProgMatchScore",
            "match_method": "ProgMatchMethod",
            "OldOddzialNazwa": "ProgMatchOldClass",
            "threshold_label": "ProgMatchLabel",
        }
    )
    classes = classes.drop(
        columns=[
            "Prog_min_klasa",
            "threshold_year",
            "threshold_kind",
            "threshold_label",
            "ProgMatchStatus",
            "ProgMatchScore",
            "ProgMatchMethod",
            "ProgMatchOldClass",
            "ProgMatchLabel",
        ],
        errors="ignore",
    ).merge(selected, on="source_class_id", how="left")
    classes["ProgMatchStatus"] = classes["ProgMatchStatus"].fillna("school_only")
    classes["ProgMatchMethod"] = classes["ProgMatchMethod"].fillna("")
    classes["ProgMatchOldClass"] = classes["ProgMatchOldClass"].fillna("")
    classes["ProgMatchLabel"] = classes["ProgMatchLabel"].fillna("")
    return classes


def add_threshold_usage_labels(df_classes: pd.DataFrame) -> pd.DataFrame:
    classes = df_classes.copy()
    classes["ProgUsedLevel"] = "brak progu"
    trusted = (
        classes["ProgMatchStatus"].eq("trusted") & classes["Prog_min_klasa"].notna()
    )
    approximate = (
        classes["ProgMatchStatus"].eq("approximate") & classes["Prog_min_klasa"].notna()
    )
    school_only = classes["Prog_min_klasa"].isna() & classes["Prog_min_szkola"].notna()
    class_years = (
        classes["threshold_year"]
        if "threshold_year" in classes.columns
        else pd.Series(pd.NA, index=classes.index)
    )
    school_years = (
        classes["Prog_szkola_threshold_year"]
        if "Prog_szkola_threshold_year" in classes.columns
        else pd.Series(pd.NA, index=classes.index)
    )
    classes.loc[trusted, "ProgUsedLevel"] = class_years[trusted].apply(
        lambda value: (
            f"klasowy {format_threshold_year(value)} - dokładny"
            if format_threshold_year(value)
            else "klasowy - dokładny"
        )
    )
    classes.loc[approximate, "ProgUsedLevel"] = class_years[approximate].apply(
        lambda value: (
            f"klasowy {format_threshold_year(value)} - przybliżony"
            if format_threshold_year(value)
            else "klasowy - przybliżony"
        )
    )
    classes.loc[school_only, "ProgUsedLevel"] = school_years[school_only].apply(
        lambda value: (
            f"szkolny {format_threshold_year(value)} - brak dopasowania klasy"
            if format_threshold_year(value)
            else "szkolny - brak dopasowania klasy"
        )
    )
    return classes


def build_school_details(df_schools: pd.DataFrame) -> pd.DataFrame:
    if df_schools.empty:
        return pd.DataFrame()
    details = df_schools.copy()
    details["OpisSzkolyMarkdown"] = (
        details["OpisSzkolyText"].fillna("")
        if "OpisSzkolyText" in details.columns
        else ""
    )
    details["OpisSzkolyPreview"] = details["OpisSzkolyMarkdown"].apply(preview_text)
    columns = [
        "source_school_id",
        "SzkolaIdentyfikator",
        "NazwaSzkoly",
        "AdresSzkoly",
        "Dzielnica",
        "Telefon",
        "Email",
        "WWW",
        "OfertaPzoUrl",
        "OpisSzkolyPreview",
        "OpisSzkolyMarkdown",
        "year",
        "admission_year",
        "school_year",
    ]
    return details[[col for col in columns if col in details.columns]].copy()


def build_class_details(
    df_classes: pd.DataFrame, criteria_summary: pd.DataFrame
) -> pd.DataFrame:
    if df_classes.empty:
        return pd.DataFrame()
    details = df_classes.copy()
    if not criteria_summary.empty:
        details = details.merge(criteria_summary, on="source_class_id", how="left")
    details["OpisOddzialuMarkdown"] = (
        details["OpisOddzialuText"].fillna("")
        if "OpisOddzialuText" in details.columns
        else ""
    )
    details["OpisOddzialuPreview"] = details["OpisOddzialuMarkdown"].apply(
        lambda value: preview_text(value, limit=450)
    )
    columns = [
        "source_class_id",
        "source_school_id",
        "SzkolaIdentyfikator",
        "NazwaSzkoly",
        "OddzialNazwa",
        "OddzialKod",
        "TypOddzialu",
        "LiczbaMiejsc",
        "LiczbaOddzialow",
        "PierwszyJezykObcy",
        "DrugiJezykObcy",
        "JezykiObce",
        "JezykiObceIkonyOpis",
        "JezykiPierwszeNorm",
        "JezykiDrugieNorm",
        "JezykiWszystkieNorm",
        "JezykiPierwszePoziomy",
        "JezykiDrugiePoziomy",
        "JezykiWszystkiePoziomy",
        "JezykiPierwszeOpcje",
        "JezykiDrugieOpcje",
        "JezykiWszystkieOpcje",
        "PrzedmiotyRozszerzone",
        "Zawod",
        "DyscyplinaSportowa",
        "Punktowany1",
        "Punktowany2",
        "Punktowany3",
        "Punktowany4",
        "PrzedmiotyPunktowane",
        "KryteriaPunktowane",
        "OpisOddzialuPreview",
        "OpisOddzialuMarkdown",
        "Prog_min_klasa",
        "Prog_min_szkola",
        "ProgMatchStatus",
        "ProgMatchScore",
        "ProgMatchMethod",
        "ProgMatchOldClass",
        "ProgCandidatesSummary",
        "ProgUsedLevel",
        "WWW",
        "OfertaPzoUrl",
        "year",
        "admission_year",
        "school_year",
    ]
    return details[[col for col in columns if col in details.columns]].copy()


def attach_pzo_cached_travel_time(
    df_schools: pd.DataFrame, location_cache: pd.DataFrame
) -> pd.DataFrame:
    schools = df_schools.copy()
    if location_cache.empty or "SzkolaIdentyfikator" not in location_cache.columns:
        schools["CzasDojazdu"] = None
        return schools
    cols = [
        col
        for col in ["SzkolaIdentyfikator", "CzasDojazdu"]
        if col in location_cache.columns
    ]
    if len(cols) == 2:
        schools = schools.drop(columns=["CzasDojazdu"], errors="ignore").merge(
            location_cache[cols]
            .dropna(subset=["SzkolaIdentyfikator"])
            .drop_duplicates("SzkolaIdentyfikator"),
            on="SzkolaIdentyfikator",
            how="left",
        )
    if "CzasDojazdu" not in schools.columns:
        schools["CzasDojazdu"] = None
    return schools


def build_vulcan_year(
    year_cfg: dict[str, Any], cfg: dict[str, Any], location_cache: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    df_vulcan = prepare_vulcan_offer(load_vulcan_offer(year_cfg), cfg)
    df_thresholds = load_thresholds(year_cfg)
    df_ranking = load_ranking(year_cfg)
    class_thresholds = best_thresholds_for_keys(
        df_thresholds,
        ["SzkolaIdentyfikator", "OddzialNazwa"],
    )

    df_classes = pd.merge(
        df_vulcan,
        class_thresholds[
            [
                "OddzialNazwa",
                "Prog_min_klasa",
                "SzkolaIdentyfikator",
                "threshold_year",
                "threshold_kind",
                "threshold_label",
            ]
        ],
        how="left",
        on=["SzkolaIdentyfikator", "OddzialNazwa"],
    )
    df_classes = pd.merge(
        df_classes,
        df_ranking[["RankingPoz", "RankingPozTekst", "SzkolaIdentyfikator"]],
        how="left",
        on="SzkolaIdentyfikator",
    )
    df_classes.drop(columns=["NazwaSzkoly_y"], errors="ignore", inplace=True)
    if "NazwaSzkoly_x" in df_classes.columns:
        df_classes.rename(columns={"NazwaSzkoly_x": "NazwaSzkoly"}, inplace=True)
    df_classes = _append_duplicate_suffix(df_classes)

    minmax = school_threshold_summary(df_thresholds)
    historical_thresholds = historical_school_thresholds(df_thresholds)
    df_schools = df_vulcan.drop_duplicates(
        subset=["SzkolaIdentyfikator", "NazwaSzkoly", "AdresSzkoly", "TypSzkoly"]
    )[
        [
            "SzkolaIdentyfikator",
            "source_school_id",
            "NazwaSzkoly",
            "AdresSzkoly",
            "TypSzkoly",
            "IdSzkoly",
            "Dzielnica",
        ]
    ].copy()
    df_schools = df_schools.merge(
        df_ranking[["SzkolaIdentyfikator", "RankingPoz", "RankingPozTekst"]],
        how="left",
        on="SzkolaIdentyfikator",
    )
    df_schools = df_schools.merge(minmax, how="left", on="SzkolaIdentyfikator")
    df_schools = df_schools.merge(
        historical_thresholds, how="left", on="SzkolaIdentyfikator"
    )
    df_schools = attach_location_data(df_schools, cfg, location_cache)
    df_schools["url"] = (
        "https://warszawa.edu.com.pl/kandydat/app/offer_school_details.xhtml?schoolId="
        + df_schools["IdSzkoly"].astype(str)
    )

    school_metric_cols = [
        "CzasDojazdu",
        "SzkolaLat",
        "SzkolaLon",
        "Prog_min_szkola",
        "Prog_max_szkola",
        "Prog_szkola_threshold_year",
        "Prog_szkola_threshold_kind",
        "Prog_szkola_threshold_label",
        "Progi_historyczne_szkola",
        "Progi_historyczne_lata",
    ]
    school_metric_keys = ["SzkolaIdentyfikator"]
    if (
        "source_school_id" in df_classes.columns
        and "source_school_id" in df_schools.columns
    ):
        school_metric_keys = ["source_school_id", "SzkolaIdentyfikator"]
    school_metrics = df_schools[
        school_metric_keys + school_metric_cols
    ].drop_duplicates(school_metric_keys)
    df_classes = df_classes.merge(
        school_metrics,
        how="left",
        on=school_metric_keys,
    )
    df_classes = add_common_class_columns(df_classes)

    threshold_info = threshold_meta(year_cfg)
    for df in [df_schools, df_classes, df_thresholds, df_ranking]:
        add_year_metadata(df, year_cfg, threshold_info)

    return {
        "schools": df_schools,
        "classes": df_classes,
        "thresholds": df_thresholds,
        "ranking": df_ranking,
        "plan": pd.DataFrame(),
    }


def _append_duplicate_suffix(df: pd.DataFrame) -> pd.DataFrame:
    school_key = "SzkolaIdentyfikator"
    if "source_school_id" in df.columns and df["source_school_id"].notna().any():
        school_key = "source_school_id"
    key = [school_key, "OddzialNazwa"]
    duplicate_no = df.groupby(key).cumcount()
    mask = duplicate_no > 0
    df.loc[mask, "OddzialNazwa"] = (
        df.loc[mask, "OddzialNazwa"] + " #" + (duplicate_no[mask] + 1).astype(str)
    )
    return df


def build_pzo_year(
    year_cfg: dict[str, Any], cfg: dict[str, Any], location_cache: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    _ = cfg
    pzo_tables = load_pzo_offer_tables(year_cfg)
    df_schools = pzo_tables.get("schools", pd.DataFrame()).copy()
    df_classes = pzo_tables.get("classes", pd.DataFrame()).copy()
    criteria_long = pzo_tables.get("criteria_long", pd.DataFrame()).copy()
    if df_schools.empty or df_classes.empty:
        raise ValueError("Snapshot PZO nie zawiera wymaganych tabel schools/classes.")

    df_thresholds = load_thresholds(year_cfg)
    df_ranking = load_ranking(year_cfg)

    reference_schools = reference_schools_from_cache(location_cache)
    df_schools, df_classes = attach_stable_school_ids(
        df_schools, df_classes, reference_schools
    )

    df_schools["OfertaPzoUrl"] = df_schools["pzo_school_type_ids"].apply(pzo_search_url)
    df_schools["url"] = (
        df_schools.get("WWW", pd.Series(index=df_schools.index, dtype=object))
        .replace("", pd.NA)
        .combine_first(df_schools["OfertaPzoUrl"])
    )
    df_schools["OpisSzkolyPreview"] = (
        df_schools["OpisSzkolyText"].fillna("").apply(preview_text)
        if "OpisSzkolyText" in df_schools.columns
        else ""
    )
    df_schools = attach_pzo_cached_travel_time(df_schools, location_cache)

    ranking_cols = ["SzkolaIdentyfikator", "RankingPoz", "RankingPozTekst"]
    if not df_ranking.empty and set(ranking_cols).issubset(df_ranking.columns):
        df_schools = df_schools.merge(df_ranking[ranking_cols], how="left")

    threshold_matches, _ = match_reference_thresholds(df_classes, df_thresholds)
    df_classes = apply_threshold_matches(df_classes, threshold_matches)

    minmax = school_threshold_summary(df_thresholds)
    historical_thresholds = historical_school_thresholds(df_thresholds)
    df_schools = df_schools.merge(minmax, how="left", on="SzkolaIdentyfikator")
    df_schools = df_schools.merge(
        historical_thresholds, how="left", on="SzkolaIdentyfikator"
    )

    school_metric_cols = [
        "source_school_id",
        "SzkolaIdentyfikator",
        "CzasDojazdu",
        "SzkolaLat",
        "SzkolaLon",
        "Prog_min_szkola",
        "Prog_max_szkola",
        "Prog_szkola_threshold_year",
        "Prog_szkola_threshold_kind",
        "Prog_szkola_threshold_label",
        "Progi_historyczne_szkola",
        "Progi_historyczne_lata",
        "RankingPoz",
        "RankingPozTekst",
        "WWW",
        "OfertaPzoUrl",
        "url",
    ]
    school_metric_keys = ["source_school_id", "SzkolaIdentyfikator"]
    df_classes = df_classes.drop(
        columns=[
            col
            for col in school_metric_cols
            if col not in school_metric_keys and col in df_classes.columns
        ],
        errors="ignore",
    )
    df_classes = df_classes.merge(
        df_schools[
            [col for col in school_metric_cols if col in df_schools.columns]
        ].drop_duplicates("source_school_id"),
        how="left",
        on=school_metric_keys,
    )
    df_classes = add_threshold_usage_labels(df_classes)
    df_classes = add_common_class_columns(df_classes)

    criteria_summary = summarize_criteria(criteria_long)
    threshold_info = threshold_meta(year_cfg)
    frames_for_meta = [
        df_schools,
        df_classes,
        df_thresholds,
        df_ranking,
        threshold_matches,
    ]
    for df in frames_for_meta:
        if df.empty:
            continue
        add_year_metadata(df, year_cfg, threshold_info)

    school_details = build_school_details(df_schools)
    class_details = build_class_details(df_classes, criteria_summary)

    school_drop_cols = ["Dyrektor", "OpisSzkolyHtml", "OpisSzkolyText"]
    class_drop_cols = [
        "OpisOddzialuHtml",
        "OpisOddzialuText",
        "QualificationGroup",
        "QualificationGroupId",
        "ModuleId",
        "BlockApply",
        "HasCriteria",
        "ShowCriteria",
        "latitude",
        "longitude",
    ]
    df_schools = df_schools.drop(columns=school_drop_cols, errors="ignore")
    df_classes = df_classes.drop(columns=class_drop_cols, errors="ignore")

    return {
        "schools": df_schools,
        "classes": df_classes,
        "thresholds": df_thresholds,
        "ranking": df_ranking,
        "school_details": school_details,
        "class_details": class_details,
        "threshold_matches": threshold_matches,
    }


def process_year(
    year_cfg: dict[str, Any], cfg: dict[str, Any], location_cache: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    offer_type = year_cfg["offer"]["type"]
    if offer_type == "vulcan_legacy":
        return build_vulcan_year(year_cfg, cfg, location_cache)
    if offer_type == "pzo_omikron":
        return build_pzo_year(year_cfg, cfg, location_cache)
    raise ValueError(f"Nieznany typ oferty: {offer_type}")


def validate_year_data(
    year_cfg: dict[str, Any], schools: pd.DataFrame, classes: pd.DataFrame
) -> dict[str, Any]:
    duplicate_count = 0
    school_key = (
        "source_school_id"
        if "source_school_id" in classes.columns
        else "SzkolaIdentyfikator"
    )
    duplicate_columns = {"year", school_key, "OddzialNazwa"}
    if duplicate_columns.issubset(classes.columns):
        duplicate_count = classes.duplicated(
            subset=["year", school_key, "OddzialNazwa"]
        ).sum()
    classes_with_threshold = (
        int(classes["Prog_min_klasa"].notna().sum())
        if "Prog_min_klasa" in classes
        else 0
    )
    classes_with_school_threshold = (
        int(classes["Prog_min_szkola"].notna().sum())
        if "Prog_min_szkola" in classes
        else 0
    )
    schools_with_threshold = (
        int(schools["Prog_min_szkola"].notna().sum())
        if "Prog_min_szkola" in schools
        else 0
    )
    threshold_info = threshold_meta(year_cfg)
    return {
        "year": year_cfg["year"],
        "school_year": year_cfg.get("school_year"),
        "data_status": year_cfg.get("data_status"),
        "status_label": year_cfg.get("status_label"),
        "threshold_mode": threshold_info["threshold_mode"],
        "threshold_label": threshold_info["threshold_label"],
        "threshold_years": threshold_info["threshold_years"],
        "schools_count": len(schools),
        "classes_count": len(classes),
        "schools_with_threshold": schools_with_threshold,
        "classes_with_threshold": classes_with_threshold,
        "classes_with_school_threshold": classes_with_school_threshold,
        "schools_without_district": (
            int(schools["Dzielnica"].isna().sum()) if "Dzielnica" in schools else 0
        ),
        "schools_without_ranking": (
            int(schools["RankingPoz"].isna().sum()) if "RankingPoz" in schools else 0
        ),
        "classes_without_threshold": (len(classes) - classes_with_threshold),
        "duplicate_class_keys": int(duplicate_count),
    }


def build_metadata(year_configs: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for year_cfg in year_configs:
        threshold_info = threshold_meta(year_cfg)
        rows.append(
            {
                "year": year_cfg["year"],
                "admission_year": year_cfg.get("admission_year"),
                "school_year": year_cfg.get("school_year"),
                "data_status": year_cfg.get("data_status"),
                "status_label": year_cfg.get("status_label"),
                "threshold_mode": threshold_info["threshold_mode"],
                "threshold_label": threshold_info["threshold_label"],
                "threshold_years": threshold_info["threshold_years"],
                "ranking_source": year_cfg.get("ranking", {}).get("source_url"),
                "threshold_source": threshold_info["threshold_source"],
                "offer_source": year_cfg.get("offer", {}).get("source_url"),
                "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
            }
        )
    return pd.DataFrame(rows)


def export_app_workbook(
    output_path: Path,
    datasets: list[dict[str, pd.DataFrame]],
    metadata: pd.DataFrame,
    quality: pd.DataFrame,
    replace_years: set[int] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    def concat(name: str) -> pd.DataFrame:
        frames = [
            dataset.get(name, pd.DataFrame())
            for dataset in datasets
            if not dataset.get(name, pd.DataFrame()).empty
        ]
        clean_frames = [frame.dropna(axis=1, how="all") for frame in frames]
        clean_frames = [frame for frame in clean_frames if not frame.empty]
        return (
            pd.concat(clean_frames, ignore_index=True, sort=False)
            if clean_frames
            else pd.DataFrame()
        )

    sheets = {
        "metadata": metadata,
        "quality": quality,
        "schools": concat("schools"),
        "classes": concat("classes"),
        "rankings": concat("ranking"),
        "thresholds": concat("thresholds"),
        "school_details": concat("school_details"),
        "class_details": concat("class_details"),
        "threshold_matches": concat("threshold_matches"),
    }
    if replace_years:
        sheets = merge_existing_year_sheets(
            read_app_workbook_sheets(output_path), sheets, replace_years
        )
    sheets = apply_latest_rankings(sheets)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    for ws in wb.worksheets:
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    logger.info("Zapisano plik aplikacyjny: %s", output_path)


def run_pipeline(year: int | None = None) -> Path:
    cfg = project_config()
    sources = source_config()
    years_config = sources["years"]
    selected_configs = []
    for key, value in years_config.items():
        year_value = int(value.get("year", key))
        if year is None or year_value == year:
            value = value.copy()
            value["year"] = year_value
            selected_configs.append(value)
    if not selected_configs:
        raise ValueError(f"Nie znaleziono konfiguracji dla roku {year}")

    location_cache = load_location_cache()
    datasets = []
    quality_rows = []
    for year_cfg in selected_configs:
        logger.info("Przetwarzanie roku danych %s", year_cfg["year"])
        dataset = process_year(year_cfg, cfg, location_cache)
        datasets.append(dataset)
        quality_rows.append(
            validate_year_data(year_cfg, dataset["schools"], dataset["classes"])
        )
        location_cache = load_location_cache()

    output_path = resolve_path(sources["app_data_file"])
    export_app_workbook(
        output_path=output_path,
        datasets=datasets,
        metadata=build_metadata(selected_configs),
        quality=pd.DataFrame(quality_rows),
        replace_years={year} if year is not None else None,
    )
    return output_path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Buduje dane aplikacji Licea Warszawa")
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Rok danych do zbudowania. Brak wartosci buduje wszystkie lata z konfiguracji.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> Path:
    args = parse_args(argv)
    return run_pipeline(year=args.year)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s"
    )
    main()
